"""Tests for Excel report generator."""

import pytest
from pathlib import Path
from tempfile import TemporaryDirectory

from src.models import HelpdeskRequest, SLA
from src.config import OutputConfig
from src.excel_generator import (
    sort_requests,
    request_to_row,
    ExcelReportGenerator,
    generate_report,
)


class TestSortRequests:
    """Tests for request sorting."""
    
    def test_sort_by_category(self):
        """Test sorting by category."""
        requests = [
            HelpdeskRequest(
                id="1",
                short_description="Z test",
                requester_email="test@test.com",
                request_category="Security",
                request_type="Type A",
            ),
            HelpdeskRequest(
                id="2",
                short_description="A test",
                requester_email="test@test.com",
                request_category="Access Management",
                request_type="Type A",
            ),
        ]
        sorted_reqs = sort_requests(requests)
        assert sorted_reqs[0].request_category == "Access Management"
        assert sorted_reqs[1].request_category == "Security"
    
    def test_sort_hierarchical(self):
        """Test hierarchical sorting."""
        requests = [
            HelpdeskRequest(
                id="1",
                short_description="B test",
                requester_email="test@test.com",
                request_category="Access",
                request_type="Type B",
            ),
            HelpdeskRequest(
                id="2",
                short_description="A test",
                requester_email="test@test.com",
                request_category="Access",
                request_type="Type A",
            ),
            HelpdeskRequest(
                id="3",
                short_description="A test",
                requester_email="test@test.com",
                request_category="Access",
                request_type="Type B",
            ),
        ]
        sorted_reqs = sort_requests(requests)
        
        # Same category, sorted by type then description
        assert sorted_reqs[0].request_type == "Type A"
        assert sorted_reqs[1].request_type == "Type B"
        assert sorted_reqs[1].short_description == "A test"
        assert sorted_reqs[2].short_description == "B test"


class TestRequestToRow:
    """Tests for row conversion."""
    
    def test_request_to_row(self):
        """Test converting request to row."""
        request = HelpdeskRequest(
            id="req_001",
            short_description="Test short",
            long_description="Test long",
            requester_email="test@example.com",
            request_category="Access Management",
            request_type="Password Reset",
            sla=SLA(unit="hours", value=4),
        )
        row = request_to_row(request)
        
        assert row[0] == "req_001"
        assert row[1] == "Test short"
        assert row[2] == "Test long"
        assert row[3] == "test@example.com"
        assert row[4] == "Access Management"
        assert row[5] == "Password Reset"
        assert row[6] == 4
        assert row[7] == "hours"


class TestExcelReportGenerator:
    """Tests for Excel report generation."""
    
    def test_generate_report(self):
        """Test full report generation."""
        requests = [
            HelpdeskRequest(
                id="req_001",
                short_description="Password reset",
                long_description="Forgot password",
                requester_email="user@example.com",
                request_category="Access Management",
                request_type="Reset forgotten password",
                sla=SLA(unit="hours", value=4),
            ),
            HelpdeskRequest(
                id="req_002",
                short_description="New laptop",
                long_description="Need new hardware",
                requester_email="user2@example.com",
                request_category="Hardware Support",
                request_type="Laptop Repair/Replacement",
                sla=SLA(unit="days", value=7),
            ),
        ]
        
        with TemporaryDirectory() as tmpdir:
            config = OutputConfig(
                output_dir=Path(tmpdir),
                report_filename="test_report.xlsx",
            )
            
            report_path = generate_report(requests, config)
            
            assert report_path.exists()
            assert report_path.suffix == ".xlsx"
            
            # Verify file is valid Excel
            from openpyxl import load_workbook
            wb = load_workbook(report_path)
            ws = wb.active
            
            # Check headers
            assert ws.cell(1, 1).value == "Request ID"
            assert ws.cell(1, 5).value == "Category"
            
            # Check data rows (sorted: Access Management before Hardware Support)
            assert ws.cell(2, 1).value == "req_001"
            assert ws.cell(3, 1).value == "req_002"
            
            # Check header styling
            assert ws.cell(1, 1).font.bold is True

