"""
Excel report generator for the Ticket Automation System.

Generates formatted Microsoft Excel reports with:
- Bold headers
- Auto-fitted columns
- Hierarchical sorting
"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import HelpdeskRequest
from .config import OutputConfig


logger = logging.getLogger(__name__)


class ExcelGeneratorError(Exception):
    """Error during Excel generation."""
    pass


# Define column configuration
COLUMN_CONFIG = [
    {"key": "id", "header": "Request ID", "width": 12},
    {"key": "short_description", "header": "Short Description", "width": 40},
    {"key": "long_description", "header": "Long Description", "width": 60},
    {"key": "requester_email", "header": "Requester Email", "width": 30},
    {"key": "request_category", "header": "Category", "width": 25},
    {"key": "request_type", "header": "Request Type", "width": 35},
    {"key": "sla_value", "header": "SLA Value", "width": 12},
    {"key": "sla_unit", "header": "SLA Unit", "width": 12},
]


def sort_requests(requests: list[HelpdeskRequest]) -> list[HelpdeskRequest]:
    """
    Sort requests hierarchically by category, type, and description.
    
    Sorting order (all ascending, standard ASCII/lexicographic):
    1. request_category
    2. request_type
    3. short_description
    
    Note: Uses standard ascending sort (case-sensitive) as per task requirements.
    
    Args:
        requests: List of requests to sort.
        
    Returns:
        Sorted list of requests.
    """
    return sorted(
        requests,
        key=lambda r: (
            r.request_category,
            r.request_type,
            r.short_description,
        )
    )


def request_to_row(request: HelpdeskRequest) -> list[Any]:
    """
    Convert a HelpdeskRequest to a row of values.
    
    Args:
        request: The request to convert.
        
    Returns:
        List of cell values matching COLUMN_CONFIG order.
    """
    return [
        request.id,
        request.short_description,
        request.long_description,
        request.requester_email,
        request.request_category,
        request.request_type,
        request.sla.value,
        request.sla.unit,
    ]


class ExcelReportGenerator:
    """
    Generator for formatted Excel reports.
    
    Produces professional-looking reports with:
    - Styled headers (bold, colored background)
    - Auto-fitted column widths
    - Text wrapping for long content
    - Proper borders
    """
    
    # Style configuration
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    CELL_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    
    # Alternating row colors for readability
    ROW_FILL_ODD = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    def __init__(self, config: OutputConfig):
        """
        Initialize the generator.
        
        Args:
            config: Output configuration with file paths.
        """
        self._config = config
    
    def generate(self, requests: list[HelpdeskRequest]) -> Path:
        """
        Generate an Excel report from classified requests.
        
        Args:
            requests: List of classified helpdesk requests.
            
        Returns:
            Path to the generated Excel file.
            
        Raises:
            ExcelGeneratorError: If report generation fails.
        """
        try:
            # Sort requests hierarchically
            sorted_requests = sort_requests(requests)
            logger.info(f"Sorted {len(sorted_requests)} requests for report")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Classified Tickets"
            
            # Write headers
            self._write_headers(ws)
            
            # Write data rows
            self._write_data(ws, sorted_requests)
            
            # Apply column widths
            self._apply_column_widths(ws)
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Ensure output directory exists
            self._config.output_dir.mkdir(parents=True, exist_ok=True)
            
            # Save workbook
            output_path = self._config.report_path
            wb.save(output_path)
            
            logger.info(f"Excel report saved to: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            raise ExcelGeneratorError(f"Report generation failed: {e}") from e
    
    def _write_headers(self, ws: Worksheet) -> None:
        """Write and style header row."""
        for col_idx, col_config in enumerate(COLUMN_CONFIG, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config["header"])
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.CELL_BORDER
        
        # Set header row height
        ws.row_dimensions[1].height = 30
    
    def _write_data(self, ws: Worksheet, requests: list[HelpdeskRequest]) -> None:
        """Write data rows with styling."""
        for row_idx, request in enumerate(requests, 2):
            row_data = request_to_row(request)
            
            # Determine row fill (alternating colors)
            fill = self.ROW_FILL_ODD if row_idx % 2 == 0 else self.ROW_FILL_EVEN
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.CELL_BORDER
                cell.fill = fill
    
    def _apply_column_widths(self, ws: Worksheet) -> None:
        """Apply column widths from configuration."""
        for col_idx, col_config in enumerate(COLUMN_CONFIG, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = col_config["width"]


def generate_report(
    requests: list[HelpdeskRequest], 
    config: OutputConfig
) -> Path:
    """
    Convenience function to generate an Excel report.
    
    Args:
        requests: List of classified requests.
        config: Output configuration.
        
    Returns:
        Path to generated report.
    """
    generator = ExcelReportGenerator(config)
    return generator.generate(requests)

